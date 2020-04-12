import sys
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from xlsx2html import xlsx2html
import pdfkit

wb = openpyxl.load_workbook(filename = '/home/olga/Загрузки/Schet-na-oplatu-82-ot-01-iyulya-2016-g-obrazets.xlsx')
sheet=wb['TDSheet']
val = sheet['B4'].value
print(val,'-',end='')
sheet['B2'].value=input()

val=sheet['B5'].value
print(val,'-',end='')
sheet['E5'].value=input()

val=sheet['M5'].value
print(val,'-',end='')
sheet['O5'].value=input()

val=sheet['X2'].value
print(val,'-',end='')
sheet['AD2'].value=input()

val=sheet['X3'].value
print(val,'-',end='')
sheet['AD3'].value=input()

val=sheet['B8'].value
print(val,'-',end='')
sheet['B6'].value=input()

val=sheet['X5'].value
print(val,'-',end='')
sheet['AD5'].value=input()

val=sheet['B14'].value
print(val,'-',end='')
sheet['G14'].value=input()

val=sheet['B17'].value
print(val,'-',end='')
sheet['G17'].value=input()

val=sheet['B20'].value
print(val,'-',end='')
sheet['B20'].value=input()

val=sheet['D22'].value
val1=sheet['B22'].value
print('Введите количество товаров/услуг-',end='')
count=int(input())
buf=[]
for i in range(count+1):
    if i>=1:
        buf+=str(i)
        buf+='<br>'
    elif i==count:
        buf+=str(i)        
stroka=''.join(buf)

sheet['B23'].value=stroka

buf=[]
stroka=''
for i in range(count):
    print('Товар/услуга ',i+1,'-',end='')
    stroka=input()
    buf+=stroka
    buf+='<br>'
stroka=''.join(buf)
sheet['D23'].value=stroka

val=sheet['AF22'].value
buf1=[]
stroka=''
buf2=[]
itogo=0
summa=0
kolvo=0
buf3=[]
for i in range(count):
    print('Кол-во ',i+1,'услуг/товаров-',end='')
    stroka=input()
    kolvo=int(stroka)
    buf1+=stroka
    buf1+='<br>'
    print(val,' ',i+1,'услуи/товара-',end='')
    stroka=input()
    summa=kolvo*int(stroka)
    buf2+=stroka
    buf2+='<br>'
    buf3+=str(summa)
    buf3+='<br>'
    itogo+=summa
stroka=''.join(buf1)
sheet['Z23'].value=stroka
stroka=''.join(buf2)
sheet['AF23'].value=stroka
stroka=''.join(buf3)
sheet['AK23'].value=stroka


val=sheet['B10'].value
print("Счет на оплату № ...  от ...20__ г. -",end='')
stroka=input()
sheet['B10'].value=stroka

val=sheet['AK25'].value
sheet['AL25'].value=itogo

val=sheet['AK26'].value
print(val,'-',end='')
sheet.merge_cells('AD26:AK26')
sheet['AD26'].value=val
HDC=int(input())
sheet['AL26'].value=HDC

val=sheet['AK27'].value
sheet.merge_cells('AD27:AK27')
sheet['AD27'].value=val
sheet['AL27'].value=HDC+itogo

itogo+=HDC

buf=[]
buf+="Всего наименований, на сумму "
buf+=str(itogo)
buf+="руб."
stroka=''.join(buf)
sheet['B28'].value=stroka

val=sheet['B29'].value
print("Итого=",itogo)
print("прим. Два миллиона пятьсот тысяч рублей 00 копеек -",end='')
stroka=input()
sheet['B29'].value=stroka

val=sheet['B37'].value
print(val,'-',end='')
sheet['M37'].value=input()

val=sheet['Z37'].value
print(val,'-',end='')
sheet['AJ37'].value=input()

wb.save('newtab.xlsx')

xlsx2html('newtab.xlsx', 'page.html')
pdfkit.from_file('page.html','Blak.pdf')


